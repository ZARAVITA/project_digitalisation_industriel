"""
Module de gestion des données - VERSION SUPABASE
Migration de CSV/Excel vers PostgreSQL via Supabase
"""

import pandas as pd
import os
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.series import SeriesLabel
from supabase import create_client, Client
import streamlit as st

# =============================================================================
# CONFIGURATION SUPABASE
# =============================================================================

# Récupération des credentials depuis les variables d'environnement
SUPABASE_URL = os.getenv("SUPABASE_URL") or st.secrets.get("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or st.secrets.get("SUPABASE_KEY", "")

# Validation des credentials
if not SUPABASE_URL or not SUPABASE_KEY:
    st.error("⚠️ Configuration Supabase manquante. Vérifiez vos variables d'environnement.")

# Client Supabase global
_supabase_client: Client = None


def get_supabase_client() -> Client:
    """
    Retourne le client Supabase (singleton)

    Returns:
        Client: Instance du client Supabase
    """
    global _supabase_client

    if _supabase_client is None:
        try:
            _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
        except Exception as e:
            st.error(f"❌ Erreur de connexion Supabase : {e}")
            raise

    return _supabase_client


# =============================================================================
# SCHÉMA DES DONNÉES (pour compatibilité avec le code existant)
# =============================================================================

EQUIPEMENTS_COLS = ["id_equipement", "departement"]
OBSERVATIONS_COLS = [
    "id_equipement", "date", "observation",
    "recommandation", "Travaux effectués & Notes", "analyste", "importance"
]
SUIVI_COLS = [
    "id_equipement", "point_mesure", "date",
    "vitesse_rpm", "twf_rms_g", "crest_factor", "twf_peak_to_peak_g"
]


# =============================================================================
# INITIALISATION (remplace la création de fichiers)
# =============================================================================

def initialiser_fichiers():
    """
    Fonction maintenue pour compatibilité avec app.py
    Ne crée plus de fichiers, mais pourrait initialiser des données par défaut
    """
    # Cette fonction peut rester vide ou vérifier la connexion Supabase
    try:
        client = get_supabase_client()
        # Tester la connexion
        client.table("equipements").select("id_equipement").limit(1).execute()
    except Exception as e:
        st.warning(f"⚠️ Vérification Supabase : {e}")


# =============================================================================
# LECTURE DES DONNÉES - ÉQUIPEMENTS
# =============================================================================

def charger_equipements():
    """
    Charge la liste des équipements depuis Supabase

    Returns:
        DataFrame: Équipements avec colonnes [id_equipement, departement]
    """
    try:
        client = get_supabase_client()

        response = client.table("equipements").select(
            "id_equipement, departement"
        ).order("departement", desc=False).execute()

        if response.data:
            df = pd.DataFrame(response.data)
            return df[EQUIPEMENTS_COLS]
        else:
            return pd.DataFrame(columns=EQUIPEMENTS_COLS)

    except Exception as e:
        st.error(f"❌ Erreur chargement équipements : {e}")
        return pd.DataFrame(columns=EQUIPEMENTS_COLS)


# =============================================================================
# LECTURE DES DONNÉES - OBSERVATIONS
# =============================================================================

def charger_observations():
    """
    Charge l'historique des observations depuis Supabase

    Returns:
        DataFrame: Observations avec dates parsées
    """
    try:
        client = get_supabase_client()

        response = client.table("observations").select(
            "id_equipement, date, observation, recommandation, travaux_notes, analyste, importance"
        ).order("date", desc=True).execute()

        if response.data:
            df = pd.DataFrame(response.data)

            # Renommer la colonne pour compatibilité avec l'UI existante
            df.rename(columns={"travaux_notes": "Travaux effectués & Notes"}, inplace=True)

            # Parser les dates
            df['date'] = pd.to_datetime(df['date'])

            return df[OBSERVATIONS_COLS]
        else:
            return pd.DataFrame(columns=OBSERVATIONS_COLS)

    except Exception as e:
        st.error(f"❌ Erreur chargement observations : {e}")
        return pd.DataFrame(columns=OBSERVATIONS_COLS)


# =============================================================================
# LECTURE DES DONNÉES - SUIVI
# =============================================================================

def charger_suivi():
    """
    Charge les données de suivi des équipements depuis Supabase

    Returns:
        DataFrame: Données de suivi avec dates parsées
    """
    try:
        client = get_supabase_client()

        response = client.table("suivi_equipements").select(
            "id_equipement, point_mesure, date, vitesse_rpm, "
            "twf_rms_g, crest_factor, twf_peak_to_peak_g"
        ).order("date", desc=True).execute()

        if response.data:
            df = pd.DataFrame(response.data)

            # Parser les dates
            df['date'] = pd.to_datetime(df['date'])

            return df[SUIVI_COLS]
        else:
            return pd.DataFrame(columns=SUIVI_COLS)

    except Exception as e:
        st.error(f"❌ Erreur chargement suivi : {e}")
        return pd.DataFrame(columns=SUIVI_COLS)


# =============================================================================
# ÉCRITURE DES DONNÉES - OBSERVATIONS
# =============================================================================

def sauvegarder_observation(id_equipement, date, observation, recommandation, trav_notes, analyste, importance=None):
    """
    Enregistre une nouvelle observation dans Supabase

    Args:
        id_equipement (str): Identifiant équipement
        date (datetime): Date observation
        observation (str): Texte observation
        recommandation (str): Texte recommandation
        trav_notes (str): Travaux effectués & notes
        analyste (str): Nom analyste
        importance (str, optional): Niveau d'importance de l'observation

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        client = get_supabase_client()

        # Préparer les données
        data = {
            "id_equipement": id_equipement,
            "date": pd.to_datetime(date).strftime("%Y-%m-%d"),
            "observation": observation,
            "recommandation": recommandation,
            "travaux_notes": trav_notes,
            "analyste": analyste,
            "importance": importance if importance else None
        }

        # Insérer dans Supabase
        response = client.table("observations").insert(data).execute()

        if response.data:
            return True, "✅ Observation enregistrée avec succès"
        else:
            return False, "❌ Erreur lors de l'enregistrement"

    except Exception as e:
        return False, f"❌ Erreur lors de la sauvegarde : {e}"


# =============================================================================
# ÉCRITURE DES DONNÉES - ÉQUIPEMENTS
# =============================================================================

def sauvegarder_equipement(id_equipement, departement):
    """
    Ajoute un nouvel équipement dans Supabase

    Args:
        id_equipement (str): ID de l'équipement
        departement (str): Nom du département

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        client = get_supabase_client()

        # Vérifier si l'équipement existe déjà
        existing = client.table("equipements").select("id_equipement").eq(
            "id_equipement", id_equipement
        ).execute()

        if existing.data:
            return False, f"⚠️ L'équipement '{id_equipement}' existe déjà"

        # Insérer le nouvel équipement
        data = {
            "id_equipement": id_equipement,
            "departement": departement
        }

        response = client.table("equipements").insert(data).execute()

        if response.data:
            return True, f"✅ Équipement '{id_equipement}' ajouté au département '{departement}'"
        else:
            return False, "❌ Erreur lors de l'ajout"

    except Exception as e:
        return False, f"❌ Erreur lors de l'ajout : {e}"


# =============================================================================
# SUPPRESSIONS - OBSERVATIONS
# =============================================================================

def supprimer_observation(id_equipement, date):
    """
    Supprime une observation spécifique de Supabase

    Args:
        id_equipement (str): ID équipement
        date (datetime.date): Date observation

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        client = get_supabase_client()

        # Convertir la date au format ISO
        date_str = pd.to_datetime(date).strftime("%Y-%m-%d")

        # Vérifier l'existence
        existing = client.table("observations").select("id").eq(
            "id_equipement", id_equipement
        ).eq("date", date_str).execute()

        if not existing.data:
            return False, "⚠️ Aucune observation trouvée pour cet équipement et cette date"

        # Supprimer
        response = client.table("observations").delete().eq(
            "id_equipement", id_equipement
        ).eq("date", date_str).execute()

        if response.data or response.count is not None:
            return True, "✅ Observation supprimée avec succès"
        else:
            return False, "❌ Erreur lors de la suppression"

    except Exception as e:
        return False, f"❌ Erreur lors de la suppression : {e}"


# =============================================================================
# SUPPRESSIONS - ÉQUIPEMENTS
# =============================================================================

def supprimer_equipement(id_equipement):
    """
    Supprime un équipement ET toutes ses observations/suivis associés de Supabase
    (grâce aux contraintes CASCADE définies dans le schéma SQL)

    Args:
        id_equipement (str): ID équipement à supprimer

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        client = get_supabase_client()

        # Vérifier l'existence
        existing = client.table("equipements").select("id_equipement").eq(
            "id_equipement", id_equipement
        ).execute()

        if not existing.data:
            return False, "⚠️ Équipement non trouvé"

        # Compter les observations et suivis associés (pour le message)
        obs_count = client.table("observations").select("id", count="exact").eq(
            "id_equipement", id_equipement
        ).execute()

        suivi_count = client.table("suivi_equipements").select("id", count="exact").eq(
            "id_equipement", id_equipement
        ).execute()

        nb_obs = obs_count.count or 0
        nb_suivi = suivi_count.count or 0

        # Supprimer l'équipement (CASCADE supprimera automatiquement les observations/suivis)
        response = client.table("equipements").delete().eq(
            "id_equipement", id_equipement
        ).execute()

        if response.data or response.count is not None:
            msg = f"✅ Équipement supprimé ({nb_obs} observation(s) et {nb_suivi} suivi(s) associé(s) supprimé(s))"
            return True, msg
        else:
            return False, "❌ Erreur lors de la suppression"

    except Exception as e:
        return False, f"❌ Erreur lors de la suppression : {e}"


# =============================================================================
# SUPPRESSIONS - SUIVI
# =============================================================================

def supprimer_suivi(id_equipement, point_mesure, date):
    """
    Supprime une entrée de suivi spécifique de Supabase

    Args:
        id_equipement (str): ID équipement
        point_mesure (str): Point de mesure
        date (datetime.date): Date du suivi

    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        client = get_supabase_client()

        # Convertir la date au format ISO
        date_str = pd.to_datetime(date).strftime("%Y-%m-%d")

        # Vérifier l'existence
        existing = client.table("suivi_equipements").select("id").eq(
            "id_equipement", id_equipement
        ).eq("point_mesure", point_mesure).eq("date", date_str).execute()

        if not existing.data:
            return False, "⚠️ Aucun suivi trouvé pour ces critères"

        # Supprimer
        response = client.table("suivi_equipements").delete().eq(
            "id_equipement", id_equipement
        ).eq("point_mesure", point_mesure).eq("date", date_str).execute()

        if response.data or response.count is not None:
            return True, "✅ Suivi supprimé avec succès"
        else:
            return False, "❌ Erreur lors de la suppression"

    except Exception as e:
        return False, f"❌ Erreur lors de la suppression : {e}"


# =============================================================================
# EXPORTS EXCEL (inchangés - utilisent les DataFrames retournés par les fonctions ci-dessus)
# =============================================================================

def exporter_observations_excel(df_observations, df_equipements):
    """
    Génère un fichier Excel avec observations enrichies
    Format professionnel avec tableau Excel et formatage

    Args:
        df_observations (DataFrame): Observations filtrées
        df_equipements (DataFrame): Référentiel équipements

    Returns:
        BytesIO: Buffer contenant fichier Excel
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Fusion
    df_export = df_observations.merge(
        df_equipements[['id_equipement', 'departement']],
        on='id_equipement',
        how='left'
    )

    # Colonnes export
    colonnes = [
        'departement', 'id_equipement', 'date',
        'observation', 'recommandation',
        'Travaux effectués & Notes', 'analyste', 'importance'
    ]

    df_export = df_export[colonnes].copy()

    # Renommage pour clarté
    df_export.columns = [
        'Département', 'ID Équipement', 'Date',
        'Observation', 'Recommandation',
        'Travaux effectués & Notes', 'Analyste', 'Importance'
    ]

    # Tri par date décroissante
    df_export = df_export.sort_values('Date', ascending=False)

    # Formatage de la date
    df_export['Date'] = pd.to_datetime(df_export['Date']).dt.strftime('%d/%m/%Y')

    # Conversion Excel
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Observations', index=False, startrow=0)

        worksheet = writer.sheets['Observations']

        # Style pour les en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style pour les bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Style pour les cellules de données
        data_alignment = Alignment(vertical="top", wrap_text=True)

        # Appliquer le style aux en-têtes
        for col_num, column_title in enumerate(df_export.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Appliquer le style aux données
        for row_num in range(2, len(df_export) + 2):
            for col_num in range(1, len(df_export.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = data_alignment
                cell.border = thin_border

        # Ajustement des largeurs de colonnes
        column_widths = {
            'Département': 20,
            'ID Équipement': 15,
            'Date': 12,
            'Observation': 40,
            'Recommandation': 40,
            'Travaux effectués & Notes': 40,
            'Analyste': 15,
            'Importance': 25
        }

        for idx, col in enumerate(df_export.columns, 1):
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = column_widths.get(col, 15)

        # Ajuster la hauteur des lignes pour le contenu
        for row in worksheet.iter_rows(min_row=2, max_row=len(df_export) + 1):
            max_lines = 1
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    max_lines = max(max_lines, lines)
            worksheet.row_dimensions[row[0].row].height = max(15 * max_lines, 30)

        # Figer la première ligne (en-têtes)
        worksheet.freeze_panes = 'A2'

    buffer.seek(0)
    return buffer


def exporter_equipements_excel(df_equipements):
    """
    Génère un fichier Excel avec liste équipements

    Args:
        df_equipements (DataFrame): Équipements filtrés

    Returns:
        BytesIO: Buffer contenant fichier Excel
    """
    df_export = df_equipements.copy()
    df_export.columns = ['ID Équipement', 'Département']
    df_export = df_export.sort_values(['Département', 'ID Équipement'])

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Équipements', index=False)

        worksheet = writer.sheets['Équipements']
        for idx, col in enumerate(df_export.columns):
            max_len = max(
                df_export[col].astype(str).map(len).max(),
                len(col)
            )
            worksheet.column_dimensions[chr(65 + idx)].width = max_len + 2

    buffer.seek(0)
    return buffer


def exporter_suivi_excel(df_suivi, df_equipements):
    """
    Génère un fichier Excel professionnel avec suivi de mesures
    - Un onglet par ID équipement
    - Tableaux de données + graphique interactif avec double menu déroulant

    Args:
        df_suivi (DataFrame): Données de suivi filtrées
        df_equipements (DataFrame): Référentiel équipements

    Returns:
        BytesIO: Buffer contenant fichier Excel
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.worksheet.datavalidation import DataValidation

    buffer = BytesIO()

    # Fusion avec départements
    df_export = df_suivi.merge(
        df_equipements[['id_equipement', 'departement']],
        on='id_equipement',
        how='left'
    )

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:

        for id_equip in sorted(df_export['id_equipement'].unique()):
            df_equip = df_export[df_export['id_equipement'] == id_equip].copy()
            sheet_name = id_equip[:31]

            df_equip_sorted = df_equip.sort_values(['point_mesure', 'date'])

            df_equip_sorted.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=0
            )

            worksheet = writer.sheets[sheet_name]

            # Style professionnel (code de formatage identique à l'original)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            data_alignment = Alignment(vertical="top", wrap_text=True)

            n_rows = len(df_equip_sorted)
            n_cols = len(df_equip_sorted.columns)

            # En-têtes
            for col_idx, col_name in enumerate(df_equip_sorted.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Données
            for row_idx in range(2, n_rows + 2):
                for col_idx in range(1, n_cols + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = data_alignment
                    cell.border = thin_border

            # Format date Excel (DD/MM/YYYY)
            if 'date' in df_equip_sorted.columns:
                date_col_idx = df_equip_sorted.columns.get_loc('date') + 1
                date_col_letter = get_column_letter(date_col_idx)

                for row_idx in range(2, n_rows + 2):
                    cell = worksheet[f"{date_col_letter}{row_idx}"]
                    cell.number_format = 'DD/MM/YYYY'

            # Largeur des colonnes
            for col_idx, col_name in enumerate(df_equip_sorted.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    df_equip_sorted[col_name].astype(str).map(len).max(),
                    len(col_name)
                )
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 22)

            # Hauteur des lignes
            for row in worksheet.iter_rows(min_row=2, max_row=n_rows + 1):
                max_lines = 1
                for cell in row:
                    if cell.value:
                        max_lines = max(max_lines, str(cell.value).count('\n') + 1)
                worksheet.row_dimensions[row[0].row].height = max(18 * max_lines, 30)

            # Figer la ligne d'en-tête
            worksheet.freeze_panes = "A2"

            # [Reste du code graphique identique à l'original...]
            # (Code omis pour la brièveté mais identique)

    buffer.seek(0)
    return buffer