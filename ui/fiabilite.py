"""
Onglet Fiabilité — Module d'analyse industrielle de la fiabilité des équipements
Calcul MTBF, taux de défaillance, courbes R(t), statistiques descriptives

Structure :
    render()
        └── render_filtres_globaux()          ← 3 filtres : Dept | Équip | Point mesure
        └── st.tabs([MTBF | Tendances | Stats])
                ├── render_tab_mtbf()         ← intervalles + MTBF + R(t) + dashboard
                ├── render_tab_tendances()    ← tous les paramètres + filtre plage + bouton détails
                └── render_tab_stats()        ← stats descriptives par paramètre

Changements v4 :
  - Filtre "Paramètre" SUPPRIMÉ des filtres globaux (inutile pour MTBF)
  - render_tab_tendances() affiche TOUS les paramètres disponibles en graphiques empilés
  - Filtre plage de valeurs : choisir le paramètre de référence + min/max
    → filtre les lignes du DataFrame, s'applique à TOUS les graphiques de tendances
  - Bouton "📊 Analyse détaillée" sous chaque courbe de tendance
    → affiche histogramme + boxplot + KDE sur TOUTES les données (sans filtre plage)
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from datetime import datetime, date
from data.data_manager import charger_equipements, charger_suivi
import io


# =============================================================================
# CONSTANTES
# =============================================================================

VARIABLES_DISPONIBLES = {
    "vitesse_rpm":        "Vitesse (RPM)",
    "twf_rms_g":          "TWF RMS (g)",
    "crest_factor":       "Crest Factor",
    "twf_peak_to_peak_g": "TWF Peak-to-Peak (g)",
}

# Palette de couleurs par paramètre pour cohérence visuelle
COULEURS_PARAMETRES = {
    "vitesse_rpm":        "#2980b9",   # bleu
    "twf_rms_g":          "#e74c3c",   # rouge
    "crest_factor":       "#f39c12",   # orange
    "twf_peak_to_peak_g": "#27ae60",   # vert
}


# =============================================================================
# UTILITAIRES — CALCULS DE FIABILITÉ
# =============================================================================

def calculer_duree_jours(debut: date, fin: date) -> float:
    """Retourne la durée en jours entre deux dates (toujours ≥ 0)."""
    return max((fin - debut).days, 0)


def calculer_fiabilite(intervalles: list) -> dict:
    """
    Calcule les indicateurs de fiabilité à partir d'une liste d'intervalles.

    Args:
        intervalles : liste de {"debut": date, "fin": date}

    Returns:
        dict avec MTBF (jours/heures), lambda, nombre_pannes, temps_total,
        durees_jours, erreur (None si tout ok)
    """
    if not intervalles:
        return None

    durees_jours      = [calculer_duree_jours(iv["debut"], iv["fin"]) for iv in intervalles]
    temps_total_jours  = sum(durees_jours)
    nombre_pannes      = max(len(intervalles) - 1, 0)

    if nombre_pannes == 0 or temps_total_jours == 0:
        return {
            "temps_total_jours":  temps_total_jours,
            "temps_total_heures": temps_total_jours * 24,
            "nombre_pannes":      nombre_pannes,
            "mtbf_jours":         None,
            "mtbf_heures":        None,
            "lambda":             None,
            "durees_jours":       durees_jours,
            "erreur": (
                "Pas assez de pannes pour calculer le MTBF "
                "(minimum 2 intervalles requis)"
            ),
        }

    mtbf_jours  = temps_total_jours / nombre_pannes
    mtbf_heures = mtbf_jours * 24
    lam         = 1.0 / mtbf_heures   # pannes / heure

    return {
        "temps_total_jours":  temps_total_jours,
        "temps_total_heures": temps_total_jours * 24,
        "nombre_pannes":      nombre_pannes,
        "mtbf_jours":         mtbf_jours,
        "mtbf_heures":        mtbf_heures,
        "lambda":             lam,
        "durees_jours":       durees_jours,
        "erreur":             None,
    }


def fiabilite_rt(lam: float, t_heures: float) -> float:
    """R(t) = exp(−λ · t)  avec λ en pannes/heure et t en heures."""
    return float(np.exp(-lam * t_heures))


def couleur_fiabilite(r: float) -> str:
    """Indicateur couleur 🟢/🟡/🔴 selon le niveau de fiabilité R(t)."""
    if r >= 0.80:
        return "🟢"
    elif r >= 0.50:
        return "🟡"
    return "🔴"


# =============================================================================
# UTILITAIRES — DATE MIN ÉQUIPEMENT
# =============================================================================

def _get_date_min_equipement(df_suivi: pd.DataFrame, id_equip: str) -> date:
    """
    Retourne la date de la première mesure disponible pour l'équipement
    (= borne minimale pour la saisie des intervalles MTBF).
    """
    df_e  = df_suivi[df_suivi["id_equipement"] == id_equip]
    dates = pd.to_datetime(df_e["date"], errors="coerce").dropna()
    return dates.min().date() if not dates.empty else date(2000, 1, 1)


# =============================================================================
# FILTRES GLOBAUX — 3 FILTRES SEULEMENT (Dept | Équip | Point de mesure)
# Le paramètre N'EST PLUS un filtre global.
# Résultats stockés dans session_state pour partage entre les 3 onglets.
# =============================================================================

def render_filtres_globaux(df_equipements: pd.DataFrame, df_suivi: pd.DataFrame):
    """
    Affiche les 3 filtres en cascade et stocke dans session_state :
        fiab_departement   → département sélectionné
        fiab_equipement    → ID équipement sélectionné
        fiab_point_mesure  → point de mesure sélectionné
        fiab_df_filtered   → DataFrame filtré (toutes colonnes, toutes variables)
        fiab_selection_ok  → True si des données existent pour la sélection
    """
    with st.container(border=True):
        st.markdown("#### 🔍 Sélection de l'équipement")
        col1, col2, col3 = st.columns(3)

        # ── 1. Département ────────────────────────────────────────────────────
        with col1:
            departements = sorted(df_equipements["departement"].dropna().unique())
            dept_idx     = 0
            if st.session_state.get("fiab_departement") in departements:
                dept_idx = departements.index(st.session_state["fiab_departement"])
            dept = st.selectbox(
                "1️⃣ Département", departements,
                index=dept_idx, key="fiab_departement"
            )

        # ── 2. ID Équipement ──────────────────────────────────────────────────
        equips_dept    = df_equipements[
            df_equipements["departement"] == dept
        ]["id_equipement"].tolist()
        equips_valides = sorted(
            [e for e in equips_dept if e in df_suivi["id_equipement"].unique()]
        )

        with col2:
            if not equips_valides:
                st.selectbox("2️⃣ ID Équipement", ["— aucun —"], key="fiab_equipement")
                st.session_state["fiab_selection_ok"] = False
                return
            eq_idx = 0
            if st.session_state.get("fiab_equipement") in equips_valides:
                eq_idx = equips_valides.index(st.session_state["fiab_equipement"])
            id_equip = st.selectbox(
                "2️⃣ ID Équipement", equips_valides,
                index=eq_idx, key="fiab_equipement"
            )

        # ── 3. Point de mesure ────────────────────────────────────────────────
        df_equip = df_suivi[df_suivi["id_equipement"] == id_equip]
        points   = sorted(df_equip["point_mesure"].dropna().unique())

        with col3:
            if not points:
                st.selectbox("3️⃣ Point de mesure", ["— aucun —"], key="fiab_point_mesure")
                st.session_state["fiab_selection_ok"] = False
                return
            pm_idx = 0
            if st.session_state.get("fiab_point_mesure") in points:
                pm_idx = points.index(st.session_state["fiab_point_mesure"])
            point_mesure = st.selectbox(
                "3️⃣ Point de mesure", points,
                index=pm_idx, key="fiab_point_mesure"
            )

    # ── DataFrame filtré (toutes les colonnes — toutes variables) ─────────────
    df_filtered = df_equip[df_equip["point_mesure"] == point_mesure].copy()
    df_filtered["date"] = pd.to_datetime(df_filtered["date"], errors="coerce")
    df_filtered = df_filtered.sort_values("date")

    # Garder seulement les colonnes qui existent réellement dans df_suivi
    cols_variables = [c for c in VARIABLES_DISPONIBLES.keys() if c in df_filtered.columns]

    st.session_state["fiab_df_filtered"]   = df_filtered
    st.session_state["fiab_cols_variables"] = cols_variables   # paramètres disponibles
    st.session_state["fiab_selection_ok"]  = not df_filtered.empty

    if df_filtered.empty:
        st.warning("⚠️ Aucune donnée pour cette sélection.")


# =============================================================================
# SECTION — INTERVALLES DE FONCTIONNEMENT
# =============================================================================

def render_intervalles(date_min: date) -> list:
    """
    Gère l'ajout, l'affichage et la suppression des intervalles MTBF.
    Clés session_state : 'fiab_intervalles', 'fiab_iv_equip_courant'.
    Réinitialise automatiquement si l'équipement change.

    Args:
        date_min : première mesure de l'équipement → borne minimale des date_input.
    """
    key_list  = "fiab_intervalles"
    key_equip = "fiab_iv_equip_courant"

    if key_list not in st.session_state:
        st.session_state[key_list] = []

    # Réinitialisation si l'équipement a changé
    equip_courant = st.session_state.get("fiab_equipement", "")
    if st.session_state.get(key_equip) != equip_courant:
        st.session_state[key_list]  = []
        st.session_state[key_equip] = equip_courant

    intervalles = st.session_state[key_list]

    st.caption(
        f"📅 Première mesure disponible pour cet équipement : "
        f"**{date_min.strftime('%d/%m/%Y')}** — "
        f"les dates ne peuvent pas être antérieures à cette date."
    )

    st.markdown("#### ➕ Ajouter un intervalle de bon fonctionnement")

    col_d1, col_d2, col_add = st.columns([2, 2, 1])
    with col_d1:
        debut_new = st.date_input(
            "Date début",
            value=date_min,
            min_value=date_min,
            max_value=date.today(),
            key="fiab_iv_debut_new"
        )
    with col_d2:
        fin_new = st.date_input(
            "Date fin",
            value=date.today(),
            min_value=date_min,
            max_value=date.today(),
            key="fiab_iv_fin_new"
        )
    with col_add:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ Ajouter", key="fiab_iv_btn_add",
                     use_container_width=True, type="primary"):
            if debut_new < date_min:
                st.error(
                    f"❌ Date de début ({debut_new.strftime('%d/%m/%Y')}) antérieure "
                    f"à la première mesure ({date_min.strftime('%d/%m/%Y')})."
                )
            elif fin_new <= debut_new:
                st.error("❌ La date de fin doit être postérieure à la date de début.")
            else:
                chevauchement = any(
                    not (fin_new <= iv["debut"] or debut_new >= iv["fin"])
                    for iv in intervalles
                )
                if chevauchement:
                    st.warning("⚠️ Cet intervalle chevauche un intervalle existant.")
                else:
                    intervalles.append({"debut": debut_new, "fin": fin_new})
                    intervalles.sort(key=lambda x: x["debut"])
                    st.session_state[key_list] = intervalles
                    st.rerun()

    if intervalles:
        st.markdown("#### 📋 Liste des intervalles")
        rows = []
        for i, iv in enumerate(intervalles):
            d = calculer_duree_jours(iv["debut"], iv["fin"])
            rows.append({
                "#":              i + 1,
                "Date début":     iv["debut"].strftime("%d/%m/%Y"),
                "Date fin":       iv["fin"].strftime("%d/%m/%Y"),
                "Durée (jours)":  d,
                "Durée (mois)":   round(d / 30.44, 1),
                "Durée (années)": round(d / 365.25, 2),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        col_sup, col_clear = st.columns([3, 1])
        with col_sup:
            idx_sup = st.selectbox(
                "Supprimer l'intervalle n°",
                options=list(range(1, len(intervalles) + 1)),
                key="fiab_iv_idx_sup"
            )
            if st.button("🗑️ Supprimer cet intervalle", key="fiab_iv_btn_sup"):
                intervalles.pop(idx_sup - 1)
                st.session_state[key_list] = intervalles
                st.rerun()
        with col_clear:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔄 Tout effacer", key="fiab_iv_btn_clear",
                         type="secondary", use_container_width=True):
                st.session_state[key_list] = []
                st.rerun()
    else:
        st.info("ℹ️ Aucun intervalle ajouté. "
                "Ajoutez au moins 2 intervalles pour calculer le MTBF.")

    return intervalles


# =============================================================================
# SECTION — KPI CARDS
# =============================================================================

def render_kpi_cards(resultats: dict, t_heures: float):
    """Affiche les 6 KPI dans des cartes HTML stylisées."""
    mtbf_h  = resultats.get("mtbf_heures")
    mtbf_j  = resultats.get("mtbf_jours")
    lam     = resultats.get("lambda")
    n_pan   = resultats.get("nombre_pannes", 0)
    t_tot_j = resultats.get("temps_total_jours", 0)
    r_t     = fiabilite_rt(lam, t_heures) if lam else None

    col1, col2, col3, col4, col5, col6 = st.columns(6)

    def card(col, titre, valeur, unite, icone, couleur="#1f77b4"):
        with col:
            st.markdown(f"""
            <div style="
                background:linear-gradient(135deg,{couleur}15 0%,{couleur}08 100%);
                border:1.5px solid {couleur}40;border-radius:12px;
                padding:16px 12px;text-align:center;min-height:110px;">
              <div style="font-size:1.6rem;margin-bottom:4px;">{icone}</div>
              <div style="font-size:0.72rem;color:#888;font-weight:600;
                text-transform:uppercase;letter-spacing:0.05em;">{titre}</div>
              <div style="font-size:1.4rem;font-weight:800;color:{couleur};
                margin:4px 0;">{valeur}</div>
              <div style="font-size:0.72rem;color:#aaa;">{unite}</div>
            </div>""", unsafe_allow_html=True)

    mc = ("#2ecc71" if mtbf_j and mtbf_j >= 365
          else ("#f39c12" if mtbf_j and mtbf_j >= 90 else "#e74c3c"))
    rc = ("#2ecc71" if r_t and r_t >= 0.80
          else ("#f39c12" if r_t and r_t >= 0.50 else "#e74c3c"))

    card(col1, "MTBF",        f"{mtbf_j:.1f}" if mtbf_j else "—", "jours",          "⚙️", mc)
    card(col2, "MTBF",        f"{mtbf_h:.0f}" if mtbf_h else "—", "heures",          "🕐", mc)
    card(col3, "Taux λ",      f"{lam:.2e}"    if lam    else "—", "pannes/heure",    "📉", "#9b59b6")
    card(col4, f"R({t_heures:.0f}h)",
         f"{r_t*100:.1f}%" if r_t is not None else "—",
         f"t = {t_heures:.0f}h", "🎯", rc)
    card(col5, "Pannes",      str(n_pan),                           "défaillances",  "⚠️", "#e74c3c")
    card(col6, "Temps total", f"{t_tot_j:.0f}",                    "jours fonct.",   "📅", "#3498db")


# =============================================================================
# SECTION — COURBE R(t)
# =============================================================================

def render_courbe_fiabilite(lam: float, mtbf_h: float):
    """Courbe R(t) = exp(−λt) avec zones colorées et repère MTBF."""
    t_max  = mtbf_h * 3
    t_vals = np.linspace(0, t_max, 500)
    r_vals = np.exp(-lam * t_vals)

    fig = go.Figure()
    fig.add_hrect(y0=0.80, y1=1.0,  fillcolor="rgba(46,204,113,0.1)",  line_width=0,
                  annotation_text="Bon (>80%)",      annotation_position="right")
    fig.add_hrect(y0=0.50, y1=0.80, fillcolor="rgba(243,156,18,0.1)",  line_width=0,
                  annotation_text="Attention",       annotation_position="right")
    fig.add_hrect(y0=0,    y1=0.50, fillcolor="rgba(231,76,60,0.1)",   line_width=0,
                  annotation_text="Critique (<50%)", annotation_position="right")
    fig.add_trace(go.Scatter(
        x=t_vals, y=r_vals, mode="lines", name="R(t) = e^(−λt)",
        line=dict(color="#2980b9", width=3),
        fill="tozeroy", fillcolor="rgba(41,128,185,0.1)"
    ))
    fig.add_vline(x=mtbf_h, line_dash="dash", line_color="#e74c3c",
                  annotation_text=f"MTBF = {mtbf_h:.0f}h (R≈36.8%)",
                  annotation_position="top right")
    fig.add_hline(y=float(np.exp(-1)), line_dash="dot",
                  line_color="#e74c3c", line_width=1)
    fig.update_layout(
        title="📈 Courbe de fiabilité R(t) — Loi exponentielle",
        xaxis_title="Temps t (heures)",
        yaxis_title="Fiabilité R(t)",
        yaxis=dict(range=[0, 1.05], tickformat=".0%"),
        hovermode="x unified", height=420,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    st.plotly_chart(fig, use_container_width=True, key="fiab_fig_rt")


# =============================================================================
# SECTION — GRAPHIQUE TENDANCE D'UN PARAMÈTRE (utilisé par Visualisation)
# =============================================================================

def _fig_tendance(df_plot: pd.DataFrame, parametre: str,
                  param_label: str, id_equip: str, point_mesure: str) -> go.Figure:
    """
    Construit et retourne la figure Plotly d'évolution temporelle pour
    un paramètre donné : données brutes + tendance linéaire + moyenne mobile 5 pts.
    """
    couleur = COULEURS_PARAMETRES.get(parametre, "#2980b9")
    fig     = go.Figure()

    fig.add_trace(go.Scatter(
        x=df_plot["date"], y=df_plot[parametre],
        mode="lines+markers", name=param_label,
        line=dict(color=couleur, width=2), marker=dict(size=5)
    ))

    if len(df_plot) >= 3:
        x_num  = (df_plot["date"] - df_plot["date"].min()).dt.days.values
        coeffs = np.polyfit(x_num, df_plot[parametre].values, 1)
        fig.add_trace(go.Scatter(
            x=df_plot["date"], y=np.polyval(coeffs, x_num),
            mode="lines", name="Tendance linéaire",
            line=dict(color="#e74c3c", width=2, dash="dash")
        ))

    if len(df_plot) >= 5:
        fig.add_trace(go.Scatter(
            x=df_plot["date"],
            y=df_plot[parametre].rolling(5, center=True).mean(),
            mode="lines", name="Moyenne mobile (5 pts)",
            line=dict(color="#2ecc71", width=2, dash="dot")
        ))

    fig.update_layout(
        title=f"📊 {param_label} — {id_equip} | {point_mesure}",
        xaxis_title="Date", yaxis_title=param_label,
        hovermode="x unified", height=380,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    return fig


# =============================================================================
# SECTION — ANALYSE DÉTAILLÉE (histogramme + boxplot + KDE + stats)
# IMPORTANT : utilise TOUTES les données (sans filtre de plage)
# =============================================================================

def render_analyse_detaillee(df_complet: pd.DataFrame, parametre: str,
                              param_label: str, prefix: str):
    """
    Affiche l'analyse statistique complète d'un paramètre sur TOUTES les données
    disponibles (sans filtre de plage de valeurs).

    Args:
        df_complet : DataFrame COMPLET de l'équipement/point de mesure
                     (sans aucun filtre de plage appliqué)
        parametre  : clé de la colonne (ex. "vitesse_rpm")
        param_label: label humain (ex. "Vitesse (RPM)")
        prefix     : préfixe unique pour les keys Plotly (ex. "vitesse_rpm")
    """
    serie = df_complet[parametre].dropna()
    if serie.empty:
        st.warning(f"⚠️ Aucune donnée pour {param_label}.")
        return

    # ── Tableau de statistiques ───────────────────────────────────────────────
    rms_val = float(np.sqrt(np.mean(serie ** 2)))
    stats   = {
        "Moyenne":           float(serie.mean()),
        "Médiane":           float(serie.median()),
        "Min":               float(serie.min()),
        "Max":               float(serie.max()),
        "Écart-type (σ)":    float(serie.std()),
        "Variance (σ²)":     float(serie.var()),
        "RMS":               rms_val,
        "P25 (Q1)":          float(serie.quantile(0.25)),
        "P75 (Q3)":          float(serie.quantile(0.75)),
        "Nombre de mesures": len(serie),
    }
    col_stat, col_box = st.columns([1, 1])

    with col_stat:
        st.markdown(f"**Statistiques — {param_label}** *(toutes données)*")
        df_s = pd.DataFrame([
            {"Indicateur": k,
             "Valeur": f"{v:.4f}" if isinstance(v, float) else str(v)}
            for k, v in stats.items()
        ])
        st.dataframe(df_s, use_container_width=True, hide_index=True)

    with col_box:
        fig_box = go.Figure()
        fig_box.add_trace(go.Box(
            y=serie, boxpoints="outliers",
            marker_color="#9b59b6", line_color="#8e44ad",
            name=param_label, fillcolor="rgba(155,89,182,0.2)"
        ))
        fig_box.update_layout(
            title=f"📦 Boxplot — {param_label}",
            yaxis_title=param_label, height=320, showlegend=False
        )
        st.plotly_chart(fig_box, use_container_width=True,
                        key=f"fiab_detail_{prefix}_box")

    # ── Histogramme ───────────────────────────────────────────────────────────
    fig_histo = go.Figure()
    fig_histo.add_trace(go.Histogram(
        x=serie, nbinsx=30,
        marker_color=COULEURS_PARAMETRES.get(parametre, "#3498db"),
        marker_line=dict(color="#2980b9", width=1),
        opacity=0.85, name=param_label
    ))
    fig_histo.add_vline(x=float(serie.mean()), line_dash="dash", line_color="#e74c3c",
                        annotation_text=f"Moy={serie.mean():.3f}",
                        annotation_position="top right")
    fig_histo.add_vline(x=float(serie.median()), line_dash="dot", line_color="#f39c12",
                        annotation_text=f"Méd={serie.median():.3f}",
                        annotation_position="top left")
    fig_histo.update_layout(
        title=f"📊 Distribution — {param_label}",
        xaxis_title=param_label, yaxis_title="Fréquence",
        height=320, showlegend=False
    )
    st.plotly_chart(fig_histo, use_container_width=True,
                    key=f"fiab_detail_{prefix}_histo")

    # ── Densité KDE (optionnel — scipy) ──────────────────────────────────────
    try:
        from scipy.stats import gaussian_kde  # type: ignore
        kde   = gaussian_kde(serie)
        x_kde = np.linspace(float(serie.min()), float(serie.max()), 300)
        fig_kde = go.Figure()
        fig_kde.add_trace(go.Scatter(
            x=x_kde, y=kde(x_kde),
            fill="tozeroy", fillcolor="rgba(46,204,113,0.2)",
            line=dict(color="#2ecc71", width=2.5), name="Densité KDE"
        ))
        fig_kde.add_vline(x=float(serie.mean()), line_dash="dash",
                          line_color="#e74c3c",
                          annotation_text="Moyenne",
                          annotation_position="top right")
        fig_kde.update_layout(
            title=f"📈 Densité de probabilité (KDE) — {param_label}",
            xaxis_title=param_label, yaxis_title="Densité", height=320
        )
        st.plotly_chart(fig_kde, use_container_width=True,
                        key=f"fiab_detail_{prefix}_kde")
    except Exception:
        pass  # scipy absent → ignoré silencieusement


# =============================================================================
# SECTION — STATISTIQUES DESCRIPTIVES (onglet 3)
# =============================================================================

def render_statistiques(df: pd.DataFrame, parametre: str, param_label: str):
    """Tableau de statistiques descriptives du paramètre (onglet Stats)."""
    serie = df[parametre].dropna()
    if serie.empty:
        st.warning("⚠️ Aucune donnée disponible pour ce paramètre.")
        return

    rms_val = float(np.sqrt(np.mean(serie ** 2)))
    stats   = {
        "Moyenne":           float(serie.mean()),
        "Médiane":           float(serie.median()),
        "Min":               float(serie.min()),
        "Max":               float(serie.max()),
        "Écart-type (σ)":    float(serie.std()),
        "Variance (σ²)":     float(serie.var()),
        "RMS":               rms_val,
        "P10":               float(serie.quantile(0.10)),
        "P25 (Q1)":          float(serie.quantile(0.25)),
        "P75 (Q3)":          float(serie.quantile(0.75)),
        "P90":               float(serie.quantile(0.90)),
        "Plage (max−min)":   float(serie.max() - serie.min()),
        "Nombre de mesures": len(serie),
    }
    df_stats = pd.DataFrame([
        {"Indicateur": k,
         "Valeur": f"{v:.4f}" if isinstance(v, float) else str(v)}
        for k, v in stats.items()
    ])
    st.dataframe(df_stats, use_container_width=True, hide_index=True)


# =============================================================================
# SECTION — EXPORTS CSV / EXCEL  (inchangé)
# =============================================================================

def render_exports(df: pd.DataFrame, parametre: str, param_label: str,
                   resultats: dict, intervalles: list,
                   id_equip: str, point_mesure: str):
    """Boutons d'export CSV et Excel du rapport MTBF."""
    st.markdown("#### 📤 Exports")
    col_csv, col_xlsx = st.columns(2)

    with col_csv:
        # Exporte toutes les colonnes variables disponibles
        cols_exp = ["date", "id_equipement", "point_mesure"] + [
            c for c in VARIABLES_DISPONIBLES.keys() if c in df.columns
        ]
        df_exp = df[cols_exp].copy()
        df_exp["date"] = pd.to_datetime(df_exp["date"]).dt.strftime("%d/%m/%Y")
        st.download_button(
            label="📥 Export CSV (données)",
            data=df_exp.to_csv(index=False, sep=";").encode("utf-8-sig"),
            file_name=(
                f"fiabilite_{id_equip}_{parametre}_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
            ),
            mime="text/csv",
            use_container_width=True
        )

    with col_xlsx:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            buf    = io.BytesIO()
            wb     = openpyxl.Workbook()
            hfill  = PatternFill("solid", fgColor="1A5276")
            hfont  = Font(bold=True, color="FFFFFF", size=11)
            bfont  = Font(bold=True, size=11)
            border = Border(
                left=Side(style="thin"),  right=Side(style="thin"),
                top=Side(style="thin"),   bottom=Side(style="thin")
            )

            # Feuille 1 : Résumé
            ws = wb.active
            ws.title = "Résumé Fiabilité"
            ws.cell(
                row=1, column=1,
                value=f"Rapport de Fiabilité — {id_equip} | {point_mesure} | {param_label}"
            ).font = Font(bold=True, size=14, color="1A5276")
            ws.cell(row=2, column=1,
                    value=f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            ws.merge_cells("A1:D1")
            for i, (lbl, val) in enumerate([
                ("Équipement",   id_equip),
                ("Point de mesure", point_mesure),
                ("Paramètre",    param_label),
                ("Temps total (jours)",
                 round(resultats.get("temps_total_jours", 0), 1)),
                ("Temps total (heures)",
                 round(resultats.get("temps_total_heures", 0), 0)),
                ("Nombre de défaillances",
                 resultats.get("nombre_pannes", "—")),
                ("MTBF (jours)",
                 round(resultats["mtbf_jours"], 2)
                 if resultats.get("mtbf_jours") else "—"),
                ("MTBF (heures)",
                 round(resultats["mtbf_heures"], 1)
                 if resultats.get("mtbf_heures") else "—"),
                ("Taux λ (pannes/h)",
                 f"{resultats['lambda']:.4e}"
                 if resultats.get("lambda") else "—"),
            ], start=4):
                c1 = ws.cell(row=i, column=1, value=lbl)
                c2 = ws.cell(row=i, column=2, value=val)
                c1.font = bfont;  c1.border = border
                c2.border = border
            ws.column_dimensions["A"].width = 38
            ws.column_dimensions["B"].width = 22

            # Feuille 2 : Intervalles
            ws2 = wb.create_sheet("Intervalles")
            for j, h in enumerate(
                ["N°", "Date début", "Date fin",
                 "Durée (jours)", "Durée (mois)", "Durée (années)"], 1
            ):
                c = ws2.cell(row=1, column=j, value=h)
                c.fill = hfill; c.font = hfont
                c.alignment = Alignment(horizontal="center"); c.border = border
            for i, iv in enumerate(intervalles, 2):
                d = calculer_duree_jours(iv["debut"], iv["fin"])
                for j, val in enumerate(
                    [i - 1, iv["debut"].strftime("%d/%m/%Y"),
                     iv["fin"].strftime("%d/%m/%Y"),
                     d, round(d / 30.44, 1), round(d / 365.25, 2)], 1
                ):
                    c = ws2.cell(row=i, column=j, value=val)
                    c.border = border
                    c.alignment = Alignment(horizontal="center")
            for j in range(1, 7):
                ws2.column_dimensions[get_column_letter(j)].width = 16

            # Feuille 3 : Données
            ws3   = wb.create_sheet("Données")
            cols3 = ["date", "id_equipement", "point_mesure"] + [
                c for c in VARIABLES_DISPONIBLES.keys() if c in df.columns
            ]
            for j, h in enumerate(cols3, 1):
                c = ws3.cell(row=1, column=j, value=h)
                c.fill = hfill; c.font = hfont; c.border = border
            for i, (_, row) in enumerate(df[cols3].iterrows(), 2):
                ws3.cell(
                    row=i, column=1,
                    value=pd.to_datetime(row["date"]).strftime("%d/%m/%Y")
                ).border = border
                ws3.cell(row=i, column=2, value=row["id_equipement"]).border = border
                ws3.cell(row=i, column=3, value=row["point_mesure"]).border = border
                for jj, col in enumerate(cols3[3:], 4):
                    v = row[col] if col in row.index else None
                    ws3.cell(
                        row=i, column=jj,
                        value=float(v) if pd.notna(v) else None
                    ).border = border
            for j in range(1, len(cols3) + 1):
                ws3.column_dimensions[get_column_letter(j)].width = 18

            wb.save(buf); buf.seek(0)
            st.download_button(
                label="📥 Export Excel (rapport)",
                data=buf,
                file_name=(
                    f"rapport_fiabilite_{id_equip}_"
                    f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
                use_container_width=True
            )
        except ImportError:
            st.warning("⚠️ openpyxl non disponible — export Excel désactivé.")


# =============================================================================
# ONGLET 1 — CALCUL MTBF & FIABILITÉ  (inchangé dans sa logique)
# =============================================================================

def render_tab_mtbf():
    """
    Onglet MTBF & Fiabilité.
    - Lit département/équipement/point_mesure depuis session_state.
    - date_min dynamique = MIN(date) dans suivi_equipements pour l'équipement.
    - Choix d'unité pour R(t) : heures ou jours.
    - N'utilise PLUS le filtre "paramètre" (supprimé des filtres globaux).
    - Pour l'export, utilise le premier paramètre disponible comme référence.
    """
    if not st.session_state.get("fiab_selection_ok", False):
        st.info("ℹ️ Sélectionnez un équipement dans les filtres ci-dessus.")
        return

    df_filtered   = st.session_state["fiab_df_filtered"]
    id_equip      = st.session_state["fiab_equipement"]
    point_mesure  = st.session_state["fiab_point_mesure"]
    cols_variables = st.session_state.get("fiab_cols_variables", [])

    # Paramètre de référence pour l'export (premier disponible)
    param_ref   = cols_variables[0] if cols_variables else "vitesse_rpm"
    label_ref   = VARIABLES_DISPONIBLES.get(param_ref, param_ref)

    # Date min dynamique
    df_suivi_global = st.session_state.get("fiab_df_suivi_global", pd.DataFrame())
    date_min        = _get_date_min_equipement(df_suivi_global, id_equip)

    # ── Intervalles ───────────────────────────────────────────────────────────
    with st.container(border=True):
        st.subheader("📅 Intervalles de bon fonctionnement")
        st.caption(
            "Ajoutez chaque période pendant laquelle l'équipement a fonctionné "
            "correctement. Nombre de pannes = nombre d'intervalles − 1."
        )
        intervalles = render_intervalles(date_min=date_min)

    # ── Calcul MTBF ───────────────────────────────────────────────────────────
    resultats = calculer_fiabilite(intervalles) if intervalles else None

    if resultats and resultats.get("erreur"):
        st.warning(f"⚠️ {resultats['erreur']}")

    # ── Saisie t + unité pour R(t) ────────────────────────────────────────────
    with st.container(border=True):
        st.subheader("🎯 Calcul de la fiabilité R(t)")

        col_t, col_unite, col_info = st.columns([1, 1, 2])

        with col_unite:
            unite_t = st.selectbox(
                "Unité du temps",
                options=["heures", "jours"],
                index=0,
                key="fiab_rt_unite",
                help=(
                    "Si « jours » est sélectionné, "
                    "la conversion t × 24 est appliquée avant le calcul. "
                    "λ reste toujours en pannes/heure."
                )
            )

        if unite_t == "heures":
            label_t   = "Temps t (heures)"
            val_def   = 720.0
            step_t    = 24.0
            max_t     = 100_000.0
        else:
            label_t   = "Temps t (jours)"
            val_def   = 30.0
            step_t    = 1.0
            max_t     = 4_000.0

        with col_t:
            t_saisi = st.number_input(
                label_t,
                min_value=1.0, max_value=max_t,
                value=val_def, step=step_t,
                key="fiab_t_saisi",
                help="Durée pour laquelle on calcule R(t)."
            )

        t_heures        = t_saisi * 24.0 if unite_t == "jours" else t_saisi
        label_t_display = (
            f"{t_saisi:.0f} jours ({t_heures:.0f}h)"
            if unite_t == "jours"
            else f"{t_heures:.0f}h"
        )

        if resultats and not resultats.get("erreur") and resultats.get("lambda"):
            r_t        = fiabilite_rt(resultats["lambda"], t_heures)
            indicateur = couleur_fiabilite(r_t)
            with col_info:
                st.markdown(f"""
                <div style="background:#f8f9fa;border-radius:10px;
                            padding:16px;margin-top:8px;">
                  <span style="font-size:1.05rem;">
                    {indicateur}&nbsp;
                    Pour <b>t = {label_t_display}</b>, la probabilité que
                    l'équipement fonctionne sans défaillance est :
                  </span><br>
                  <span style="font-size:2rem;font-weight:800;color:#2980b9;">
                    R(t) = {r_t * 100:.2f}%
                  </span><br>
                  <span style="font-size:0.8rem;color:#888;">
                    λ = {resultats['lambda']:.4e} pannes/h —
                    exp(−{resultats['lambda']:.4e} × {t_heures:.1f})
                  </span>
                </div>
                """, unsafe_allow_html=True)

    # ── Dashboard KPI + récapitulatif + courbe R(t) ───────────────────────────
    if resultats and not resultats.get("erreur"):
        with st.container(border=True):
            st.subheader("📊 Tableau de bord — Indicateurs de fiabilité")
            render_kpi_cards(resultats, t_heures)
            st.markdown("---")

            col_tab, col_courbe = st.columns([1, 2])
            with col_tab:
                st.markdown("##### 📋 Récapitulatif")
                lam = resultats["lambda"]
                recap = {
                    "Indicateur": [
                        "Temps total de fonctionnement",
                        "Temps total (heures)",
                        "Nombre de défaillances",
                        "MTBF",
                        "MTBF (heures)",
                        "Taux de défaillance λ",
                        "t saisi",
                        "t converti (heures)",
                        "Fiabilité R(t)",
                    ],
                    "Valeur": [
                        f"{resultats['temps_total_jours']:.1f} jours",
                        f"{resultats['temps_total_heures']:.0f} h",
                        f"{resultats['nombre_pannes']} panne(s)",
                        f"{resultats['mtbf_jours']:.1f} jours"
                        if resultats.get("mtbf_jours") else "—",
                        f"{resultats['mtbf_heures']:.1f} h"
                        if resultats.get("mtbf_heures") else "—",
                        f"{lam:.4e} pannes/h" if lam else "—",
                        f"{t_saisi:.1f} {unite_t}",
                        f"{t_heures:.1f} h",
                        f"{fiabilite_rt(lam, t_heures) * 100:.2f}%" if lam else "—",
                    ]
                }
                st.dataframe(pd.DataFrame(recap),
                             use_container_width=True, hide_index=True)

            with col_courbe:
                if resultats.get("lambda") and resultats.get("mtbf_heures"):
                    render_courbe_fiabilite(
                        resultats["lambda"], resultats["mtbf_heures"]
                    )

        # ── Exports ───────────────────────────────────────────────────────────
        with st.container(border=True):
            render_exports(
                df_filtered, param_ref, label_ref,
                resultats, intervalles, id_equip, point_mesure
            )


# =============================================================================
# ONGLET 2 — VISUALISATION DES TENDANCES  (entièrement refactorisé)
#
# Nouveau comportement :
#   1. Filtre temporel (période personnalisée OU N dernières observations)
#   2. Filtre optionnel par plage de valeurs sur un paramètre de référence
#      → appliqué sur le DataFrame → filtre TOUTES les tendances
#   3. Pour chaque paramètre disponible :
#        - graphique de tendance temporelle
#        - bouton "📊 Analyse détaillée" → affiche stats/histo/boxplot/KDE
#          sur TOUTES les données (sans filtre de plage)
# =============================================================================

def render_tab_tendances():
    """
    Onglet Visualisation des tendances.
    Affiche tous les paramètres disponibles en graphiques empilés.
    """
    if not st.session_state.get("fiab_selection_ok", False):
        st.info("ℹ️ Sélectionnez un équipement dans les filtres ci-dessus.")
        return

    df_complet    = st.session_state["fiab_df_filtered"].copy()  # données COMPLÈTES
    id_equip      = st.session_state["fiab_equipement"]
    point_mesure  = st.session_state["fiab_point_mesure"]
    cols_variables = st.session_state.get("fiab_cols_variables", [])

    if not cols_variables:
        st.warning("⚠️ Aucun paramètre vibratoire disponible pour cet équipement.")
        return

    # ── 1. Filtre temporel ────────────────────────────────────────────────────
    with st.container(border=True):
        st.subheader("⚙️ Options d'affichage")
        col_o1, col_o2, col_o3 = st.columns(3)

        with col_o1:
            mode = st.radio(
                "Mode de filtrage",
                ["Période personnalisée", "N dernières observations"],
                horizontal=True, key="fiab_tend_mode"
            )

        df_temp = df_complet.copy()

        if mode == "Période personnalisée":
            with col_o2:
                d1 = st.date_input(
                    "Date début",
                    value=df_complet["date"].min().date(),
                    min_value=df_complet["date"].min().date(),
                    max_value=df_complet["date"].max().date(),
                    key="fiab_tend_d1"
                )
            with col_o3:
                d2 = st.date_input(
                    "Date fin",
                    value=df_complet["date"].max().date(),
                    min_value=df_complet["date"].min().date(),
                    max_value=df_complet["date"].max().date(),
                    key="fiab_tend_d2"
                )
            df_temp = df_temp[
                (df_temp["date"].dt.date >= d1) &
                (df_temp["date"].dt.date <= d2)
            ]
        else:
            with col_o2:
                n_obs = st.number_input(
                    "N dernières observations",
                    min_value=5, max_value=500, value=22, step=1,
                    key="fiab_tend_n"
                )
            df_temp = df_temp.tail(int(n_obs))

        # ── 2. Filtre optionnel par plage de valeurs ───────────────────────────
        st.markdown("---")
        activer_filtre = st.toggle(
            "🔎 Filtrer par plage de valeurs "
            "(s'applique à tous les graphiques de tendances)",
            value=False, key="fiab_tend_filtre_val"
        )

        df_plot = df_temp.copy()   # sera filtré si activer_filtre = True

        if activer_filtre and not df_temp.empty:
            col_ref, col_vmin, col_vmax = st.columns(3)

            with col_ref:
                param_ref_filtre = st.selectbox(
                    "Paramètre de référence",
                    options=cols_variables,
                    format_func=lambda k: VARIABLES_DISPONIBLES.get(k, k),
                    key="fiab_tend_param_ref"
                )

            val_serie = df_temp[param_ref_filtre].dropna()
            with col_vmin:
                val_min = st.number_input(
                    "Valeur minimum",
                    value=float(val_serie.min()),
                    key="fiab_tend_vmin"
                )
            with col_vmax:
                val_max = st.number_input(
                    "Valeur maximum",
                    value=float(val_serie.max()),
                    key="fiab_tend_vmax"
                )

            # Appliquer le filtre : garder seulement les lignes
            # dont param_ref_filtre est dans [val_min, val_max]
            df_plot = df_temp[
                (df_temp[param_ref_filtre] >= val_min) &
                (df_temp[param_ref_filtre] <= val_max)
            ].copy()

            n_avant  = len(df_temp)
            n_apres  = len(df_plot)
            st.caption(
                f"🔎 Filtre actif sur **{VARIABLES_DISPONIBLES.get(param_ref_filtre, param_ref_filtre)}** "
                f"entre {val_min:.3f} et {val_max:.3f} — "
                f"{n_apres} / {n_avant} lignes conservées."
            )

    if df_plot.empty:
        st.warning("⚠️ Aucune donnée après filtrage. Ajustez les critères.")
        return

    # ── 3. Graphiques de tendances — un par paramètre ────────────────────────
    st.markdown("---")

    for parametre in cols_variables:
        param_label = VARIABLES_DISPONIBLES.get(parametre, parametre)
        df_param    = df_plot[["date", parametre]].dropna().sort_values("date")

        if df_param.empty:
            st.info(f"ℹ️ Aucune donnée pour {param_label} avec le filtre actuel.")
            continue

        with st.container(border=True):
            # Graphique de tendance
            fig = _fig_tendance(df_param, parametre, param_label, id_equip, point_mesure)
            st.plotly_chart(
                fig, use_container_width=True,
                key=f"fiab_tend_graph_{parametre}"
            )

            # Bouton "Analyse détaillée" — utilise df_complet (toutes données, sans filtre)
            key_btn     = f"fiab_tend_btn_{parametre}"
            key_visible = f"fiab_tend_show_{parametre}"

            if key_visible not in st.session_state:
                st.session_state[key_visible] = False

            col_btn, col_info = st.columns([1, 4])
            with col_btn:
                if st.button(
                    "📊 Analyse détaillée",
                    key=key_btn,
                    use_container_width=True,
                    help=(
                        "Affiche histogramme, boxplot et densité KDE "
                        "sur TOUTES les données disponibles "
                        "(sans filtre de plage de valeurs)."
                    )
                ):
                    st.session_state[key_visible] = not st.session_state[key_visible]

            with col_info:
                if st.session_state[key_visible]:
                    st.caption(
                        f"📌 Les statistiques ci-dessous utilisent **toutes les données** "
                        f"de {param_label} ({len(df_complet[parametre].dropna())} mesures), "
                        f"indépendamment du filtre de plage."
                    )

            # Panneau détaillé (affiché ou masqué selon l'état du bouton)
            if st.session_state[key_visible]:
                df_complet_param = df_complet[["date", parametre]].dropna()
                render_analyse_detaillee(
                    df_complet_param, parametre, param_label,
                    prefix=parametre   # ex. "vitesse_rpm" → keys uniques garanties
                )


# =============================================================================
# ONGLET 3 — STATISTIQUES DESCRIPTIVES  (inchangé dans sa logique)
# Ajoute maintenant un selectbox paramètre LOCAL à l'onglet
# (car le paramètre n'est plus dans les filtres globaux)
# =============================================================================

def render_tab_stats():
    """
    Onglet Statistiques descriptives.
    Permet de choisir le paramètre localement (selectbox dans l'onglet).
    """
    if not st.session_state.get("fiab_selection_ok", False):
        st.info("ℹ️ Sélectionnez un équipement dans les filtres ci-dessus.")
        return

    df_base       = st.session_state["fiab_df_filtered"].copy()
    id_equip      = st.session_state["fiab_equipement"]
    point_mesure  = st.session_state["fiab_point_mesure"]
    cols_variables = st.session_state.get("fiab_cols_variables", [])

    if not cols_variables:
        st.warning("⚠️ Aucun paramètre vibratoire disponible pour cet équipement.")
        return

    # ── Sélection locale du paramètre ────────────────────────────────────────
    with st.container(border=True):
        st.subheader("⚙️ Paramètre et période")
        col_p, col_d1, col_d2 = st.columns(3)

        with col_p:
            parametre = st.selectbox(
                "Paramètre",
                options=cols_variables,
                format_func=lambda k: VARIABLES_DISPONIBLES.get(k, k),
                key="fiab_stat_parametre"
            )
            param_label = VARIABLES_DISPONIBLES.get(parametre, parametre)

        df_base["date"] = pd.to_datetime(df_base["date"], errors="coerce")
        with col_d1:
            ds1 = st.date_input(
                "Date début",
                value=df_base["date"].min().date(),
                key="fiab_stat_d1"
            )
        with col_d2:
            ds2 = st.date_input(
                "Date fin",
                value=df_base["date"].max().date(),
                key="fiab_stat_d2"
            )

    df_stat = df_base[
        (df_base["date"].dt.date >= ds1) &
        (df_base["date"].dt.date <= ds2)
    ]

    if df_stat.empty:
        st.warning("⚠️ Aucune donnée pour cette période.")
        return

    # ── Statistiques ──────────────────────────────────────────────────────────
    with st.container(border=True):
        st.subheader(f"📋 Statistiques descriptives — {param_label}")
        render_statistiques(df_stat, parametre, param_label)

    # ── Graphiques ────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.subheader(f"📊 Visualisations — {param_label}")

        activer_filtre = st.toggle(
            "🔎 Filtrer les valeurs",
            value=False, key="fiab_stat_filtre"
        )
        val_min_s = val_max_s = None
        if activer_filtre:
            col_v1, col_v2 = st.columns(2)
            with col_v1:
                val_min_s = st.number_input(
                    "Valeur min",
                    value=float(df_stat[parametre].min()),
                    key="fiab_stat_vmin"
                )
            with col_v2:
                val_max_s = st.number_input(
                    "Valeur max",
                    value=float(df_stat[parametre].max()),
                    key="fiab_stat_vmax"
                )

        # Données à utiliser pour les graphiques de l'onglet Stats
        if val_min_s is not None and val_max_s is not None:
            df_viz = df_stat[
                (df_stat[parametre] >= val_min_s) &
                (df_stat[parametre] <= val_max_s)
            ]
        else:
            df_viz = df_stat

        if df_viz.empty:
            st.warning("⚠️ Aucune donnée après filtrage.")
            return

        # Graphique tendance
        df_param_viz = df_viz[["date", parametre]].dropna().sort_values("date")
        if not df_param_viz.empty:
            fig_t = _fig_tendance(
                df_param_viz, parametre, param_label, id_equip, point_mesure
            )
            st.plotly_chart(fig_t, use_container_width=True,
                            key="fiab_stat_fig_tend")

        # Analyse détaillée (histo + box + KDE)
        render_analyse_detaillee(
            df_viz, parametre, param_label,
            prefix="stat"
        )


# =============================================================================
# POINT D'ENTRÉE PRINCIPAL
# =============================================================================

def render():
    """
    Structure finale :

        🔧 Analyse de Fiabilité
        ┌─────────────────────────────────┐
        │  Filtres globaux (3 colonnes)   │
        │  Dept | Équipement | Pt mesure  │
        └─────────────────────────────────┘
        ┌──────────────────────────────────────────────────────┐
        │  [⚙️ Calcul MTBF | 📈 Tendances | 📊 Statistiques]  │
        └──────────────────────────────────────────────────────┘
    """
    st.header("🔧 Analyse de Fiabilité")
    st.caption(
        "Calcul du MTBF, taux de défaillance, "
        "courbes R(t) et statistiques industrielles"
    )

    df_equipements = charger_equipements()
    df_suivi       = charger_suivi()

    if df_equipements.empty or df_suivi.empty:
        st.error("⚠️ Données insuffisantes. Vérifiez la connexion à la base de données.")
        return

    # Stocker df_suivi complet pour calcul date_min dans render_tab_mtbf()
    st.session_state["fiab_df_suivi_global"] = df_suivi

    # Filtres globaux (3 colonnes — sans paramètre)
    render_filtres_globaux(df_equipements, df_suivi)

    st.markdown("---")

    # Onglets
    tab_mtbf, tab_tend, tab_stats = st.tabs([
        "⚙️ Calcul MTBF & Fiabilité",
        "📈 Visualisation des tendances",
        "📊 Statistiques descriptives",
    ])

    with tab_mtbf:
        render_tab_mtbf()

    with tab_tend:
        render_tab_tendances()

    with tab_stats:
        render_tab_stats()